import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox, filedialog
import tkinter.ttk as ttk
import pandas as pd
from datetime import datetime
import logging
import asyncio
import aiohttp
import threading
import json
import os
import sys
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from utils import (
    get_supabase_client, save_company_info, get_all_team_leaders,
    get_team_companies, get_all_companies, delete_company, update_company,
    save_download_history, get_company_download_history, is_new_company,
    setup_logging, save_to_supabase, get_existing_subvention_ids,
    get_download_history_records, get_subventions_by_ids
)
from code_map import *

# CustomTkinter 설정
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ──────────────────────────────────────────────
# 테마 & 상수
# ──────────────────────────────────────────────
THEME = {
    "primary": "#1a73e8",
    "primary_hover": "#1557b0",
    "primary_light": "#e8f0fe",
    "secondary": "#5f6368",
    "secondary_hover": "#4a4e52",
    "success": "#1e8e3e",
    "success_hover": "#188038",
    "danger": "#d93025",
    "danger_hover": "#b31412",
    "warning": "#f9ab00",
    "info": "#1a73e8",
    "bg_main": "#f1f3f4",
    "bg_card": "#ffffff",
    "bg_input": "#f8f9fa",
    "text_primary": "#202124",
    "text_secondary": "#5f6368",
    "text_light": "#80868b",
    "border": "#dadce0",
    "header_bg": "#1a73e8",
    "header_text": "#ffffff",
    "row_even": "#ffffff",
    "row_odd": "#f8f9fa",
    "selected_row": "#e8f0fe",
    "white": "#ffffff",
}

FONT_FAMILY = "Apple SD Gothic Neo" if sys.platform == "darwin" else "맑은 고딕"

COMPANY_TABLE_COLUMNS = {
    0: {"name": "팀장", "weight": 0, "minsize": 150},
    1: {"name": "업체명", "weight": 0, "minsize": 300},
    2: {"name": "업종", "weight": 0, "minsize": 300},
    3: {"name": "지역", "weight": 0, "minsize": 150},
    4: {"name": "시군구", "weight": 0, "minsize": 150},
}

SETTINGS_FILE = "app_settings.json"


def make_font(size=12, weight="normal"):
    return ctk.CTkFont(family=FONT_FAMILY, size=size, weight=weight)


def load_settings():
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"설정 로드 실패: {e}")
    return {"download_folder": os.getcwd()}


def save_settings(settings):
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"설정 저장 실패: {e}")
        return False


def parse_end_date(end_date_str):
    """다양한 형식의 마감일 문자열을 datetime으로 파싱"""
    if not end_date_str or end_date_str == '확인 필요':
        return None
    for fmt in ('%Y.%m.%d', '%y.%m.%d', '%Y-%m-%d', '%Y%m%d'):
        try:
            return datetime.strptime(end_date_str.strip(), fmt)
        except ValueError:
            continue
    return None


# ──────────────────────────────────────────────
# 크롤링 관련 함수
# ──────────────────────────────────────────────
async def scrape_detail_page(session, subvention_id, log_callback, semaphore, max_retries=5, retry_delay=3):
    async with semaphore:
        url = f"https://internal.pay.naver.com/partner/api/subvention/detail/{subvention_id}"
        retries = 0

        while retries < max_retries:
            try:
                timeout = aiohttp.ClientTimeout(total=30)
                async with session.get(url, timeout=timeout) as response:
                    if response.status == 200:
                        data = await response.json()

                        logging.debug(f"API 응답 데이터 (subvention_id={subvention_id}): {data}")

                        if not data or "data" not in data:
                            if log_callback:
                                log_callback(f"{subvention_id}의 상세 데이터가 비어 있습니다.")
                            return None

                        details = data["data"]

                        area = ", ".join(
                            [a.get("areaName", "확인 필요") or "확인 필요" for a in details.get("subventionAreaList", []) if a]
                        ) or "확인 필요"
                        institution = details.get("receptionInstitutionName", "확인 필요") or "확인 필요"
                        subvention_title = details.get("subventionTitleName", "확인 필요") or "확인 필요"
                        support_method = ", ".join(
                            [m.get("description", "확인 필요") or "확인 필요" for m in details.get("subventionSupportMethodCodeList", []) if m]
                        ) or "확인 필요"
                        support_amount = f"{details.get('supportAmount', 0):,} 원" if details.get("supportAmount") else "확인 필요"

                        reception_end_date = details.get("receptionEndYmd", "확인 필요") or "확인 필요"
                        subvention_url = details.get(
                            "subventionUrlAddress", f"https://finsupport.naver.com/subvention/detail/{subvention_id}"
                        ) or f"https://finsupport.naver.com/subvention/detail/{subvention_id}"
                        source_name = details.get("sourceName", "확인 필요") or "확인 필요"

                        lowest_interest = details.get("lowestInterest")
                        highest_interest = details.get("highestInterest")
                        if lowest_interest and highest_interest:
                            interest_rate = f"{lowest_interest}% ~ {highest_interest}%"
                        elif lowest_interest:
                            interest_rate = f"{lowest_interest}% 이상"
                        elif highest_interest:
                            interest_rate = f"{highest_interest}% 이하"
                        else:
                            interest_rate = "확인 필요"

                        application_way_html = details.get("applicationWayHtmlContent")
                        application_method = "확인 필요"
                        if application_way_html:
                            try:
                                if "<" in application_way_html and ">" in application_way_html:
                                    soup = BeautifulSoup(application_way_html, "html.parser")
                                    application_method = soup.get_text(" ", strip=True)
                                else:
                                    application_method = application_way_html.strip()
                            except Exception as e:
                                logging.warning(f"BeautifulSoup HTML 파싱 실패: {e}")
                                application_method = "확인 필요"

                        attachments = "\n".join(
                            [
                                f"{file.get('linkFileName', '파일 이름 없음')}: {file.get('linkFileUrlAddress', '링크 없음')}"
                                for file in details.get("subventionFileList", []) if file
                            ]
                        ) if details.get("appendixFileYn") == "Y" else "확인 필요"

                        business_type_code = "ALL"
                        if "businessTypeCode" in details:
                            business_type_code = details["businessTypeCode"]
                        elif "subventionBusinessTypeCodeList" in details:
                            business_codes = details["subventionBusinessTypeCodeList"]
                            if business_codes and len(business_codes) > 0:
                                business_type_code = business_codes[0].get("businessTypeCode", "ALL")

                        return {
                            "지역": area,
                            "접수기관": institution,
                            "지원사업명": subvention_title,
                            "지원 방식": support_method,
                            "지원금액": support_amount,
                            "금리": interest_rate,
                            "접수 마감일": reception_end_date,
                            "접수 방법": application_method,
                            "공고 URL": subvention_url,
                            "출처": source_name,
                            "첨부파일": attachments,
                            "subventionId": subvention_id,
                            "businessTypeCode": business_type_code,
                        }
                    elif response.status in [429, 500, 502, 503, 504]:
                        retries += 1
                        backoff_delay = retry_delay * (2 ** (retries - 1))
                        if log_callback:
                            log_callback(f"{subvention_id} 요청 중 {response.status} 오류 발생. {retries}/{max_retries}회 재시도 중... {backoff_delay}초 대기.")
                        await asyncio.sleep(backoff_delay)
                    else:
                        if log_callback:
                            log_callback(f"{subvention_id}의 상세 정보를 가져오는 데 실패했습니다. 상태 코드: {response.status}")
                        return None
            except asyncio.TimeoutError:
                retries += 1
                backoff_delay = retry_delay * (2 ** (retries - 1))
                if log_callback:
                    log_callback(f"{subvention_id} 요청 중 타임아웃 발생. {retries}/{max_retries}회 재시도 중... {backoff_delay}초 대기.")
                await asyncio.sleep(backoff_delay)
            except Exception as e:
                retries += 1
                backoff_delay = retry_delay * (2 ** (retries - 1))
                if log_callback:
                    log_callback(f"{subvention_id} 요청 중 오류 발생: {e}. {retries}/{max_retries}회 재시도 중... {backoff_delay}초 대기.")
                await asyncio.sleep(backoff_delay)

        if log_callback:
            log_callback(f"{subvention_id}의 요청이 실패하여 데이터를 가져오지 못했습니다.")
        logging.error(f"{subvention_id}의 요청 실패: 최종 재시도 후에도 데이터를 가져오지 못했습니다.")
        return None


async def fetch_subvention_ids(session, page_number, log_callback, selected_industry, selected_area, selected_district):
    page_size = 30
    main_page_url = "https://internal.pay.naver.com/partner/api/subvention/list"
    headers = {"Content-Type": "application/json"}

    business_type_code = INDUSTRY_CODE_MAP.get(selected_industry)
    area_code = AREA_CODE_MAP.get(selected_area)

    english_name = next((k for k, v in ENGLISH_AREA_CODE_MAP.items() if v == area_code), None)
    district_map_variable = f"{english_name}_DISTRICT_CODE_MAP" if english_name else None
    district_code_map = globals().get(district_map_variable, {"전체": "00000"})
    district_code = district_code_map.get(selected_district, "00000")

    params = {
        "isActive": "Y",
        "page": page_number,
        "size": page_size,
        "sort": "ACCURACY",
    }
    if business_type_code and business_type_code != "ALL":
        params["businessTypeCode"] = business_type_code
    if selected_district != "전체" and district_code != "00000":
        params["areaCode"] = district_code
    elif area_code and area_code != "00000":
        params["areaCode"] = area_code

    try:
        async with session.get(main_page_url, params=params, headers=headers) as response:
            if response.status == 200:
                data = await response.json()
                if not data or "data" not in data or not data["data"]["content"]:
                    if log_callback:
                        log_callback(f"페이지 {page_number}에서 데이터가 비어 있습니다.")
                    return []
                await asyncio.sleep(0.5)
                return [subvention['subventionId'] for subvention in data['data']['content']]
            else:
                error_message = f"페이지 {page_number}에서 데이터를 가져오는 데 실패했습니다. 상태 코드: {response.status}"
                logging.error(error_message)
                if log_callback:
                    log_callback(error_message)
                return []
    except Exception as e:
        logging.error(f"API 요청 중 오류 발생: {e}")
        if log_callback:
            log_callback(f"API 요청 중 오류 발생: {e}")
        return []


async def crawl_all_data(log_callback):
    setup_logging()

    if log_callback:
        log_callback("기존 공고 ID 목록을 확인하는 중...")
    existing_ids = get_existing_subvention_ids()
    if log_callback:
        log_callback(f"기존 공고 {len(existing_ids)}개 확인됨. 중복 제거 크롤링 시작...")

    async with aiohttp.ClientSession() as session:
        all_data_with_codes = []

        industry_codes = [key for key in INDUSTRY_CODE_MAP.keys() if key != "전체"]

        async def collect_industry_data(industry):
            if log_callback:
                log_callback(f"{industry} 업종 데이터 수집 중...")

            page_number = 0
            industry_subvention_ids = []

            while True:
                subvention_ids = await fetch_subvention_ids(session, page_number, log_callback, industry, "전체", "전체")
                if not subvention_ids:
                    break
                industry_subvention_ids.extend(subvention_ids)
                page_number += 1

            if industry_subvention_ids:
                new_ids = [id for id in industry_subvention_ids if id not in existing_ids]
                if log_callback:
                    log_callback(f"{industry}: {len(industry_subvention_ids)}개 중 {len(new_ids)}개 신규 공고 발견")

                if new_ids:
                    semaphore = asyncio.Semaphore(5)
                    tasks = [
                        scrape_detail_page(session, subvention_id, log_callback, semaphore)
                        for subvention_id in new_ids
                    ]
                    results = await asyncio.gather(*tasks)
                else:
                    results = []

                business_type_code = INDUSTRY_CODE_MAP[industry]
                industry_data = []
                for result in results:
                    if result:
                        result["businessTypeCode"] = business_type_code
                        industry_data.append(result)

                return industry_data
            return []

        batch_size = 3
        for i in range(0, len(industry_codes), batch_size):
            batch_industries = industry_codes[i:i+batch_size]
            if log_callback:
                log_callback(f"업종 배치 {i//batch_size + 1}/{(len(industry_codes)-1)//batch_size + 1} 처리 중...")

            batch_tasks = [collect_industry_data(industry) for industry in batch_industries]
            batch_results = await asyncio.gather(*batch_tasks)

            for industry_data in batch_results:
                all_data_with_codes.extend(industry_data)

            if i + batch_size < len(industry_codes):
                await asyncio.sleep(3)

        if not all_data_with_codes:
            if log_callback:
                log_callback("공고 데이터를 찾을 수 없습니다.")
            return

        successful_data = [item for item in all_data_with_codes if item is not None]
        if log_callback:
            log_callback(f"총 {len(successful_data)}개의 공고 수집 완료. 중복 제거 중...")

        unique_data = {}
        for item in successful_data:
            subvention_id = item['subventionId']
            if subvention_id not in unique_data:
                unique_data[subvention_id] = item
            else:
                existing_codes = unique_data[subvention_id]['businessTypeCode'].split(',')
                new_code = item['businessTypeCode']
                if new_code not in existing_codes:
                    existing_codes.append(new_code)
                    unique_data[subvention_id]['businessTypeCode'] = ','.join(existing_codes)

        final_data = list(unique_data.values())
        if log_callback:
            log_callback(f"중복 제거 완료: {len(final_data)}개 고유 공고. Supabase에 저장 중...")

        save_to_supabase(final_data, log_callback, "전체", "전체", "전체")

        messagebox.showinfo("완료", f"업종별 공고 수집이 완료되었습니다!\n\n{len(final_data)}개 데이터 저장")


def run_crawling(log_callback):
    try:
        if log_callback:
            log_callback("전체 공고 데이터 수집을 시작합니다...")
        asyncio.run(crawl_all_data(log_callback))
    except Exception as e:
        if log_callback:
            log_callback(f"크롤링 중 오류 발생: {e}")
        logging.error(f"크롤링 중 오류 발생: {e}")


# ──────────────────────────────────────────────
# 다운로드 & 엑셀 함수
# ──────────────────────────────────────────────
def download_data_from_db(team_leader=None, log_callback=None, download_folder=None):
    try:
        if not team_leader:
            messagebox.showwarning("경고", "팀장을 선택해주세요.")
            return

        companies_data = get_team_companies(team_leader)
        if not companies_data:
            messagebox.showinfo("알림", f"{team_leader} 팀장의 등록된 업체가 없습니다.")
            return

        industry_area_groups = {}
        for company in companies_data:
            industry = company['industry']
            area = company['area']
            company_id = company['id']

            is_new = is_new_company(company_id)
            company['is_new'] = is_new

            key = f"{industry}_{area}"
            if key not in industry_area_groups:
                industry_area_groups[key] = {
                    'industry': industry,
                    'area': area,
                    'new_companies': [],
                    'existing_companies': []
                }

            if is_new:
                industry_area_groups[key]['new_companies'].append(company)
            else:
                industry_area_groups[key]['existing_companies'].append(company)

        supabase = get_supabase_client()
        saved_files = []

        for group_key, group_data in industry_area_groups.items():
            industry = group_data['industry']
            area = group_data['area']
            new_companies = group_data['new_companies']
            existing_companies = group_data['existing_companies']

            business_type_code = INDUSTRY_CODE_MAP.get(industry, "ALL")

            query = supabase.table("Finsupport Data").select("*")

            if business_type_code != "ALL":
                query = query.or_(f"business_type_code.like.%{business_type_code}%,business_type_code.eq.ALL")

            result = query.execute()

            filtered_data = []
            for item in result.data:
                item_area = item['area'].lower()

                is_area_match = False

                national_keywords = ["전국", "전체", "대한민국", "전 지역", "전지역", "국가", "정부"]
                if any(keyword in item_area for keyword in national_keywords):
                    is_area_match = True
                elif area != "전체":
                    if area.lower() in item_area:
                        is_area_match = True
                else:
                    is_area_match = True

                if is_area_match:
                    if not any(existing['subvention_id'] == item['subvention_id'] for existing in filtered_data):
                        filtered_data.append(item)

            if filtered_data:
                today_str = datetime.today().strftime('%Y%m%d')
                today = datetime.today()

                # [Bug5 수정] 견고한 날짜 비교로 마감일 지난 공고 제외
                valid_announcements = []
                for item in filtered_data:
                    end_date_str = item.get('reception_end_date', '')
                    if not end_date_str or end_date_str == '확인 필요':
                        valid_announcements.append(item)
                        continue

                    end_date = parse_end_date(end_date_str)
                    if end_date is None or end_date >= today:
                        valid_announcements.append(item)

                if not valid_announcements:
                    continue

                # 신규 업체용 파일 생성
                if new_companies:
                    excel_data = []
                    for item in valid_announcements:
                        excel_row = {
                            "지역": item["area"],
                            "접수기관": item["institution"],
                            "지원사업명": item["subvention_title"],
                            "지원 방식": item["support_method"],
                            "지원금액": item["support_amount"],
                            "금리": item["interest_rate"],
                            "접수 마감일": item["reception_end_date"],
                            "접수 방법": item["application_method"],
                            "공고 URL": item["subvention_url"],
                            "출처": item["source_name"],
                            "첨부파일": item["attachments"],
                            "subventionId": item["subvention_id"]
                        }
                        excel_data.append(excel_row)

                    filename = f"{team_leader}_{industry}_{area}_신규업체_{today_str}.xlsx"
                    company_names = [company['company_name'] for company in new_companies]

                    save_new_data(excel_data, filename, company_names, download_folder)
                    saved_files.append(filename)

                    subvention_ids = [item['subvention_id'] for item in valid_announcements]
                    for company in new_companies:
                        save_download_history(team_leader, company['id'], company['company_name'], subvention_ids)

                # 기존 업체용 파일 생성 (오늘 다운로드한 것만 중복 제거)
                if existing_companies:
                    new_announcements_set = set()
                    new_announcements = []
                    for company in existing_companies:
                        downloaded_ids = get_company_download_history(company['id'])

                        for item in valid_announcements:
                            sid = item['subvention_id']
                            if sid not in downloaded_ids and sid not in new_announcements_set:
                                new_announcements_set.add(sid)
                                new_announcements.append(item)

                    if new_announcements:
                        excel_data = []
                        for item in new_announcements:
                            excel_row = {
                                "지역": item["area"],
                                "접수기관": item["institution"],
                                "지원사업명": item["subvention_title"],
                                "지원 방식": item["support_method"],
                                "지원금액": item["support_amount"],
                                "금리": item["interest_rate"],
                                "접수 마감일": item["reception_end_date"],
                                "접수 방법": item["application_method"],
                                "공고 URL": item["subvention_url"],
                                "출처": item["source_name"],
                                "첨부파일": item["attachments"],
                                "subventionId": item["subvention_id"]
                            }
                            excel_data.append(excel_row)

                        filename = f"{team_leader}_{industry}_{area}_기존업체_{today_str}.xlsx"
                        company_names = [company['company_name'] for company in existing_companies]

                        save_new_data(excel_data, filename, company_names, download_folder)
                        saved_files.append(filename)

                        subvention_ids = [item['subvention_id'] for item in new_announcements]
                        for company in existing_companies:
                            save_download_history(team_leader, company['id'], company['company_name'], subvention_ids)

        if saved_files:
            file_list = '\n'.join(saved_files)
            if log_callback:
                log_callback(f"{team_leader} 팀장용 맞춤 공고를 {len(saved_files)}개 파일로 저장했습니다.")

            messagebox.showinfo("완료", f"{team_leader} 팀장의 관리업체 맞춤 공고가 완료되었습니다.\n\n생성된 파일:\n{file_list}")
        else:
            messagebox.showinfo("알림", f"{team_leader} 팀장의 관리업체에 적합한 공고가 없습니다.")

    except Exception as e:
        error_msg = f"DB 데이터 다운로드 중 오류 발생: {e}"
        if log_callback:
            log_callback(error_msg)
        logging.error(error_msg)
        messagebox.showerror("오류", error_msg)


def save_new_data(new_data, base_filename, companies_list, download_folder=None):
    df_new = pd.DataFrame(new_data)

    support_title_index = df_new.columns.get_loc('지원사업명')

    for company in reversed(companies_list):
        company = company.strip()
        if company:
            df_new.insert(support_title_index + 1, f'{company}', '')

    base_folder = download_folder or os.getcwd()

    team_leader = base_filename.split('_')[0]
    today_str = base_filename.split('_')[-1].replace('.xlsx', '')
    folder_name = f"{team_leader}_{today_str}"
    folder_path = os.path.join(base_folder, folder_name)

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    full_path = os.path.join(folder_path, base_filename)
    if os.path.exists(full_path):
        timestamp = datetime.now().strftime("%H%M%S")
        name_without_ext = base_filename.replace('.xlsx', '')
        new_filename = f"{name_without_ext}_{timestamp}.xlsx"
        full_path = os.path.join(folder_path, new_filename)

        counter = 1
        while os.path.exists(full_path):
            new_filename = f"{name_without_ext}_{timestamp}_{counter}.xlsx"
            full_path = os.path.join(folder_path, new_filename)
            counter += 1

    df_new.to_excel(full_path, index=False)
    adjust_excel_formatting(full_path, companies_list)


def adjust_excel_formatting(filename, companies_list):
    wb = load_workbook(filename)
    ws = wb.active

    column_widths = {
        'A': 27,
        'B': 36,
        'C': 50,
    }

    support_title_col = 'C'
    for i, company in enumerate(companies_list):
        company = company.strip()
        if company:
            col_letter = chr(ord(support_title_col) + i + 1)
            column_widths[col_letter] = 10

            dv = DataValidation(type="list", operator="equal", formula1='"☐,☑"')
            dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")
            ws.add_data_validation(dv)

    remaining_cols = {
        'D': 15,
        'E': 18,
        'F': 13,
        'G': 17,
        'H': 50,
        'I': 50,
        'J': 34,
        'K': 100
    }

    for col, width in remaining_cols.items():
        new_col = chr(ord(col) + len([c for c in companies_list if c.strip()]))
        column_widths[new_col] = width

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 50

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment

    wb.save(filename)
    wb.close()


# ──────────────────────────────────────────────
# 메인 앱 클래스
# ──────────────────────────────────────────────
class ModernTeamManagerApp:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("팀장별공고관리 시스템 v2.0")
        self.root.geometry("1400x920")

        self.selected_team = ctk.StringVar()
        self.status_text = ctk.StringVar(value="시스템을 준비하는 중...")

        self.settings = load_settings()

        # 캐시
        self.all_companies = []
        self.company_cache_loaded = False

        self.setup_ui()
        self.load_initial_data()

    # ──────────────────────────────────────
    # UI 구성
    # ──────────────────────────────────────
    def setup_ui(self):
        main_container = ctk.CTkFrame(self.root, fg_color=THEME["bg_main"])
        main_container.pack(fill="both", expand=True)

        # 헤더
        self._build_header(main_container)

        # 콘텐츠
        content_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=24, pady=(0, 0))

        self.tabview = ctk.CTkTabview(
            content_frame, width=1340, height=700,
            segmented_button_fg_color=THEME["bg_input"],
            segmented_button_selected_color=THEME["primary"],
            segmented_button_selected_hover_color=THEME["primary_hover"],
            segmented_button_unselected_color=THEME["bg_input"],
            segmented_button_unselected_hover_color=THEME["border"],
            command=self.on_tab_change
        )
        self.tabview.pack(fill="both", expand=True)

        crawling_tab = self.tabview.add("공고검색")
        self.setup_crawling_tab(crawling_tab)

        notice_tab = self.tabview.add("공고관리")
        self.setup_main_tab(notice_tab)

        company_tab = self.tabview.add("업체관리")
        self.setup_company_management_tab(company_tab)

        # 하단 상태 바
        self._build_status_bar(main_container)

    def _build_header(self, parent):
        header = ctk.CTkFrame(parent, fg_color=THEME["header_bg"], height=64, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        inner = ctk.CTkFrame(header, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=28, pady=12)

        ctk.CTkLabel(
            inner, text="팀장별공고관리 시스템",
            font=make_font(22, "bold"), text_color=THEME["header_text"]
        ).pack(side="left")

        ctk.CTkLabel(
            inner, text="v2.0",
            font=make_font(11), text_color="#aecbfa"
        ).pack(side="left", padx=(8, 0), pady=(6, 0))

    def _build_status_bar(self, parent):
        bar = ctk.CTkFrame(parent, fg_color=THEME["bg_card"], height=36, corner_radius=0)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)

        ctk.CTkLabel(
            bar, textvariable=self.status_text,
            font=make_font(11), text_color=THEME["text_secondary"]
        ).pack(side="left", padx=24)

    # ──────────────────────────────────────
    # 탭 1: 공고검색
    # ──────────────────────────────────────
    def setup_crawling_tab(self, parent):
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=16, pady=16)

        # 안내 카드
        info_card = self._card(container)
        info_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(
            info_card, text="자동 공고 수집",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(16, 6))

        ctk.CTkLabel(
            info_card,
            text="네이버 핀서포트에서 최신 금융지원 공고를 자동으로 수집하여 데이터베이스에 저장합니다.\n수집된 데이터는 팀장별 맞춤 공고 생성에 활용됩니다.",
            font=make_font(12), text_color=THEME["text_secondary"], justify="left"
        ).pack(anchor="w", padx=20, pady=(0, 16))

        # 실행 카드
        action_card = self._card(container)
        action_card.pack(fill="x", pady=(0, 16))

        btn_frame = ctk.CTkFrame(action_card, fg_color="transparent")
        btn_frame.pack(pady=20)

        self.crawling_btn = ctk.CTkButton(
            btn_frame, text="크롤링 시작",
            font=make_font(15, "bold"), width=240, height=46,
            fg_color=THEME["primary"], hover_color=THEME["primary_hover"],
            corner_radius=8, command=self.start_crawling
        )
        self.crawling_btn.pack()

        # 진행 상황 카드
        log_card = self._card(container)
        log_card.pack(fill="both", expand=True)

        ctk.CTkLabel(
            log_card, text="진행 상황",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(16, 8))

        self.crawling_log = ctk.CTkTextbox(
            log_card, height=280, font=make_font(11),
            fg_color=THEME["bg_input"], corner_radius=6
        )
        self.crawling_log.pack(fill="both", expand=True, padx=20, pady=(0, 16))
        self.crawling_log.insert("0.0", "크롤링 준비 완료. 시작 버튼을 클릭하세요.\n")

    def start_crawling(self):
        self.crawling_btn.configure(state="disabled", text="크롤링 진행 중...")
        self.crawling_log.delete("0.0", "end")
        self.update_crawling_log("크롤링을 시작합니다...\n")

        def crawling_thread():
            try:
                run_crawling(self.update_crawling_log)
                self.root.after(0, self.crawling_complete)
            except Exception as e:
                self.root.after(0, lambda: self.crawling_error(str(e)))

        threading.Thread(target=crawling_thread, daemon=True).start()

    def update_crawling_log(self, message):
        def update_ui():
            self.crawling_log.insert("end", f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
            self.crawling_log.see("end")
        self.root.after(0, update_ui)

    def crawling_complete(self):
        self.crawling_btn.configure(state="normal", text="크롤링 시작")
        self.update_crawling_log("크롤링이 완료되었습니다!")
        self.status_text.set("크롤링 완료 - 최신 공고 데이터가 업데이트되었습니다")

    def crawling_error(self, error_msg):
        self.crawling_btn.configure(state="normal", text="크롤링 시작")
        self.update_crawling_log(f"크롤링 중 오류 발생: {error_msg}")
        self.status_text.set("크롤링 실행 중 오류가 발생했습니다")

    # ──────────────────────────────────────
    # 탭 2: 공고관리
    # ──────────────────────────────────────
    def setup_main_tab(self, parent):
        container = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=16, pady=16)

        # 조회 조건 카드
        search_card = self._card(container)
        search_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(
            search_card, text="조회 조건 설정",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(16, 10))

        cond_frame = ctk.CTkFrame(search_card, fg_color=THEME["bg_input"], corner_radius=6)
        cond_frame.pack(fill="x", padx=20, pady=(0, 16))

        row1 = ctk.CTkFrame(cond_frame, fg_color="transparent")
        row1.pack(fill="x", padx=20, pady=14)

        ctk.CTkLabel(row1, text="팀장:", font=make_font(13, "bold"),
                     text_color=THEME["text_primary"]).pack(side="left", padx=(0, 8))

        self.team_combo = ctk.CTkComboBox(
            row1, values=[], width=300, font=make_font(12),
            variable=self.selected_team, command=self.on_team_select
        )
        self.team_combo.pack(side="left", padx=(0, 24))

        ctk.CTkButton(
            row1, text="업체관리 이동", font=make_font(12, "bold"),
            width=120, height=34, fg_color=THEME["secondary"],
            hover_color=THEME["secondary_hover"], corner_radius=6,
            command=self.show_company_tab
        ).pack(side="left")

        # 관리업체 현황 카드
        company_card = self._card(container)
        company_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(
            company_card, text="관리업체 현황",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(16, 8))

        self.company_textbox = ctk.CTkTextbox(
            company_card, height=160, font=make_font(12),
            fg_color=THEME["bg_input"], corner_radius=6
        )
        self.company_textbox.pack(fill="x", padx=20, pady=(0, 16))

        # 작업 실행 카드
        action_card = self._card(container)
        action_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(
            action_card, text="작업 실행",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(16, 10))

        btn_row = ctk.CTkFrame(action_card, fg_color="transparent")
        btn_row.pack(pady=(0, 10))

        ctk.CTkButton(
            btn_row, text="맞춤 공고 다운로드",
            font=make_font(15, "bold"), width=260, height=46,
            fg_color=THEME["primary"], hover_color=THEME["primary_hover"],
            corner_radius=8, command=self.download_data
        ).pack(side="left", padx=12)

        ctk.CTkButton(
            btn_row, text="다운로드 폴더 설정",
            font=make_font(12), width=140, height=34,
            fg_color=THEME["info"], hover_color=THEME["primary_hover"],
            corner_radius=6, command=self.open_folder_settings
        ).pack(side="left", padx=8)

        current_folder = self.settings.get("download_folder", os.getcwd())
        folder_display = current_folder if len(current_folder) < 60 else "..." + current_folder[-57:]
        self.folder_display_label = ctk.CTkLabel(
            action_card, text=f"다운로드 폴더: {folder_display}",
            font=make_font(11), text_color=THEME["text_light"]
        )
        self.folder_display_label.pack(pady=(0, 16))

        # ─── 다운로드 이력 카드 ───
        history_card = self._card(container)
        history_card.pack(fill="x", pady=(0, 16))

        h_header = ctk.CTkFrame(history_card, fg_color="transparent")
        h_header.pack(fill="x", padx=20, pady=(16, 8))

        ctk.CTkLabel(
            h_header, text="다운로드 이력",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(side="left")

        ctk.CTkButton(
            h_header, text="이력 조회", font=make_font(11, "bold"),
            width=90, height=30, fg_color=THEME["info"],
            hover_color=THEME["primary_hover"], corner_radius=6,
            command=self.load_download_history
        ).pack(side="right", padx=(8, 0))

        ctk.CTkButton(
            h_header, text="선택 이력 재다운로드", font=make_font(11, "bold"),
            width=140, height=30, fg_color=THEME["success"],
            hover_color=THEME["success_hover"], corner_radius=6,
            command=self.redownload_from_history
        ).pack(side="right")

        # 이력 Treeview
        h_tree_frame = ctk.CTkFrame(history_card, fg_color=THEME["bg_card"], corner_radius=6)
        h_tree_frame.pack(fill="x", padx=20, pady=(0, 16))

        tree_style = ttk.Style()
        tree_style.configure("History.Treeview.Heading",
                             background=THEME["primary"], foreground="white",
                             font=(FONT_FAMILY, 11, "bold"), relief="flat")
        tree_style.configure("History.Treeview",
                             background="#ffffff", foreground=THEME["text_primary"],
                             fieldbackground="#ffffff", font=(FONT_FAMILY, 11), rowheight=32)
        tree_style.map("History.Treeview",
                        background=[('selected', THEME["selected_row"])],
                        foreground=[('selected', THEME["primary"])])

        self.history_tree = ttk.Treeview(
            h_tree_frame, columns=("date", "team", "industry", "area", "count"),
            show="headings", height=6, style="History.Treeview"
        )
        self.history_tree.heading("date", text="다운로드 날짜")
        self.history_tree.heading("team", text="팀장")
        self.history_tree.heading("industry", text="업종")
        self.history_tree.heading("area", text="지역")
        self.history_tree.heading("count", text="공고 수")

        self.history_tree.column("date", width=140, anchor="center")
        self.history_tree.column("team", width=120, anchor="center")
        self.history_tree.column("industry", width=250, anchor="w")
        self.history_tree.column("area", width=120, anchor="center")
        self.history_tree.column("count", width=80, anchor="center")

        h_scroll = ttk.Scrollbar(h_tree_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=h_scroll.set)
        self.history_tree.pack(side="left", fill="x", expand=True, padx=(10, 0), pady=8)
        h_scroll.pack(side="right", fill="y", pady=8)

        # 이력 데이터 저장소
        self.history_data_map = {}

    # ──────────────────────────────────────
    # 다운로드 이력 기능
    # ──────────────────────────────────────
    def load_download_history(self):
        team_leader = self.selected_team.get()
        records = get_download_history_records(team_leader=team_leader if team_leader else None)

        # download_date를 분 단위로 잘라 같은 다운로드 세션을 그룹핑
        # (같은 시점에 다운로드된 레코드들은 동일 세션으로 간주)
        grouped = {}
        for record in records:
            raw_date = record.get('download_date', '')
            # ISO 형식 "2026-03-10T14:30:25.123+00:00" → 분 단위 "2026-03-10 14:30"
            session_key_date = raw_date[:16].replace('T', ' ') if len(raw_date) >= 16 else raw_date[:10]
            display_date = raw_date[:19].replace('T', ' ') if len(raw_date) >= 19 else raw_date[:10]
            team = record.get('team_leader', '')
            company_name = record.get('company_name', '')
            key = f"{session_key_date}_{team}_{company_name}"

            if key not in grouped:
                grouped[key] = {
                    'display_date': display_date,
                    'team': team,
                    'company_name': company_name,
                    'subvention_ids': set()
                }
            grouped[key]['subvention_ids'].add(record.get('subvention_id', ''))

        # Treeview 초기화
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        self.history_data_map = {}

        for key, data in grouped.items():
            item_id = self.history_tree.insert(
                "", "end",
                values=(data['display_date'], data['team'], data['company_name'], "", len(data['subvention_ids']))
            )
            self.history_data_map[item_id] = list(data['subvention_ids'])

        count = len(grouped)
        self.status_text.set(f"다운로드 이력 {count}건 조회 완료" if count > 0 else "다운로드 이력이 없습니다")

    def redownload_from_history(self):
        selection = self.history_tree.selection()
        if not selection:
            messagebox.showwarning("경고", "재다운로드할 이력을 선택해주세요.")
            return

        item_id = selection[0]
        subvention_ids = self.history_data_map.get(item_id, [])
        if not subvention_ids:
            messagebox.showwarning("경고", "해당 이력의 공고 데이터를 찾을 수 없습니다.")
            return

        values = self.history_tree.item(item_id)['values']
        date_str = values[0]
        team = values[1]
        company_name = values[2]

        self.status_text.set("이전 공고를 재다운로드하는 중...")
        self.root.update()

        # Supabase에서 공고 데이터 조회
        data = get_subventions_by_ids(subvention_ids)
        if not data:
            messagebox.showinfo("알림", "해당 공고 데이터를 DB에서 찾을 수 없습니다.")
            return

        excel_data = []
        for item in data:
            excel_row = {
                "지역": item.get("area", ""),
                "접수기관": item.get("institution", ""),
                "지원사업명": item.get("subvention_title", ""),
                "지원 방식": item.get("support_method", ""),
                "지원금액": item.get("support_amount", ""),
                "금리": item.get("interest_rate", ""),
                "접수 마감일": item.get("reception_end_date", ""),
                "접수 방법": item.get("application_method", ""),
                "공고 URL": item.get("subvention_url", ""),
                "출처": item.get("source_name", ""),
                "첨부파일": item.get("attachments", ""),
                "subventionId": item.get("subvention_id", "")
            }
            excel_data.append(excel_row)

        today_str = datetime.today().strftime('%Y%m%d')
        filename = f"{team}_재다운로드_{company_name}_{date_str}_{today_str}.xlsx"
        download_folder = self.settings.get("download_folder", os.getcwd())

        save_new_data(excel_data, filename, [company_name], download_folder)

        messagebox.showinfo("완료", f"재다운로드가 완료되었습니다.\n\n파일: {filename}\n공고 수: {len(excel_data)}건")
        self.status_text.set(f"재다운로드 완료 - {len(excel_data)}건")

    # ──────────────────────────────────────
    # 탭 3: 업체관리
    # ──────────────────────────────────────
    def setup_company_management_tab(self, parent):
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=16, pady=16)

        # 새 업체 추가 카드
        add_card = self._card(container)
        add_card.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(
            add_card, text="새 업체 추가",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(16, 10))

        form_bg = ctk.CTkFrame(add_card, fg_color=THEME["bg_input"], corner_radius=6)
        form_bg.pack(fill="x", padx=20, pady=(0, 16))

        input_grid = ctk.CTkFrame(form_bg, fg_color="transparent")
        input_grid.pack(pady=16, padx=20)

        # Row 0: 팀장명, 업체명, 업종
        ctk.CTkLabel(input_grid, text="팀장명:", font=make_font(12, "bold"),
                     text_color=THEME["text_primary"]).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=8)
        self.company_team_name_entry = ctk.CTkEntry(
            input_grid, width=150, font=make_font(11),
            corner_radius=6, placeholder_text="팀장명 입력"
        )
        self.company_team_name_entry.grid(row=0, column=1, padx=(0, 16), pady=8)

        ctk.CTkLabel(input_grid, text="업체명:", font=make_font(12, "bold"),
                     text_color=THEME["text_primary"]).grid(row=0, column=2, sticky="w", padx=(0, 8), pady=8)
        self.company_company_name_entry = ctk.CTkEntry(
            input_grid, width=180, font=make_font(11),
            corner_radius=6, placeholder_text="업체명 입력"
        )
        self.company_company_name_entry.grid(row=0, column=3, padx=(0, 16), pady=8)

        ctk.CTkLabel(input_grid, text="업종:", font=make_font(12, "bold"),
                     text_color=THEME["text_primary"]).grid(row=0, column=4, sticky="w", padx=(0, 8), pady=8)
        self.company_industry_combo = ctk.CTkComboBox(
            input_grid, values=list(INDUSTRY_CODE_MAP.keys()), width=280,
            font=make_font(11), corner_radius=6, state="readonly"
        )
        self.company_industry_combo.grid(row=0, column=5, pady=8)
        self.company_industry_combo.set("전체")

        # Row 1: 지역, 시군구
        ctk.CTkLabel(input_grid, text="지역:", font=make_font(12, "bold"),
                     text_color=THEME["text_primary"]).grid(row=1, column=0, sticky="w", padx=(0, 8), pady=8)
        self.company_area_select = ctk.CTkComboBox(
            input_grid, values=list(AREA_CODE_MAP.keys()), width=150,
            font=make_font(11), corner_radius=6, state="readonly",
            command=self.update_district_options
        )
        self.company_area_select.grid(row=1, column=1, padx=(0, 16), pady=8)
        self.company_area_select.set("전체")

        ctk.CTkLabel(input_grid, text="시군구:", font=make_font(12, "bold"),
                     text_color=THEME["text_primary"]).grid(row=1, column=2, sticky="w", padx=(0, 8), pady=8)
        self.company_district_combo = ctk.CTkComboBox(
            input_grid, values=["전체"], width=180,
            font=make_font(11), corner_radius=6, state="readonly"
        )
        self.company_district_combo.grid(row=1, column=3, padx=(0, 16), pady=8)
        self.company_district_combo.set("전체")

        # 저장 버튼
        ctk.CTkButton(
            input_grid, text="업체 저장", font=make_font(13, "bold"),
            width=140, height=38, fg_color=THEME["success"],
            hover_color=THEME["success_hover"], corner_radius=6,
            command=self.save_company_in_tab
        ).grid(row=2, column=2, columnspan=2, pady=12)

        # 업체 목록 카드
        list_card = self._card(container)
        list_card.pack(fill="both", expand=True)

        # 헤더 행
        list_header = ctk.CTkFrame(list_card, fg_color="transparent")
        list_header.pack(fill="x", padx=20, pady=(16, 8))

        ctk.CTkLabel(
            list_header, text="등록된 업체 목록",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(side="left")

        # 검색 + 버튼 영역
        ctrl_frame = ctk.CTkFrame(list_header, fg_color="transparent")
        ctrl_frame.pack(side="right")

        ctk.CTkLabel(ctrl_frame, text="검색:", font=make_font(12, "bold"),
                     text_color=THEME["text_primary"]).pack(side="left", padx=(0, 6))
        self.search_entry = ctk.CTkEntry(
            ctrl_frame, width=200, font=make_font(11),
            corner_radius=6, placeholder_text="팀장명 또는 업체명"
        )
        self.search_entry.pack(side="left", padx=(0, 8))
        self.search_entry.bind("<KeyRelease>", self.on_search_change)

        ctk.CTkButton(
            ctrl_frame, text="검색", width=50, height=30,
            font=make_font(11), fg_color=THEME["primary"],
            hover_color=THEME["primary_hover"], corner_radius=6,
            command=self.search_companies
        ).pack(side="left", padx=(4, 2))

        ctk.CTkButton(
            ctrl_frame, text="초기화", width=56, height=30,
            font=make_font(11), fg_color=THEME["secondary"],
            hover_color=THEME["secondary_hover"], corner_radius=6,
            command=self.reset_search
        ).pack(side="left", padx=(2, 12))

        ctk.CTkButton(
            ctrl_frame, text="수정", width=56, height=30,
            font=make_font(11), fg_color=THEME["info"],
            hover_color=THEME["primary_hover"], corner_radius=6,
            command=self.edit_selected_company
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            ctrl_frame, text="삭제", width=56, height=30,
            font=make_font(11), fg_color=THEME["danger"],
            hover_color=THEME["danger_hover"], corner_radius=6,
            command=self.delete_selected_company
        ).pack(side="left", padx=4)

        # Treeview
        tree_frame = ctk.CTkFrame(list_card, fg_color=THEME["bg_card"], corner_radius=6, height=380)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 16))
        tree_frame.pack_propagate(False)

        tree_style = ttk.Style()
        tree_style.theme_use('clam')
        tree_style.configure("Company.Treeview.Heading",
                             background=THEME["primary"], foreground="white",
                             font=(FONT_FAMILY, 12, "bold"), relief="flat")
        tree_style.configure("Company.Treeview",
                             background="#ffffff", foreground=THEME["text_primary"],
                             fieldbackground="#ffffff", font=(FONT_FAMILY, 11), rowheight=35)
        tree_style.map("Company.Treeview",
                        background=[('selected', THEME["selected_row"])],
                        foreground=[('selected', THEME["primary"])])

        self.company_tree = ttk.Treeview(
            tree_frame, columns=("team", "company", "industry", "area", "district"),
            show="headings", height=12, style="Company.Treeview"
        )

        self._sort_state = {}  # 열별 정렬 상태 (True=오름차순, False=내림차순)
        for col, label in [("team", "팀장"), ("company", "업체명"), ("industry", "업종"), ("area", "지역"), ("district", "시군구")]:
            self.company_tree.heading(col, text=label, command=lambda c=col: self._sort_company_tree(c))

        self.company_tree.column("team", width=COMPANY_TABLE_COLUMNS[0]["minsize"], anchor="w")
        self.company_tree.column("company", width=COMPANY_TABLE_COLUMNS[1]["minsize"], anchor="w")
        self.company_tree.column("industry", width=COMPANY_TABLE_COLUMNS[2]["minsize"], anchor="w")
        self.company_tree.column("area", width=COMPANY_TABLE_COLUMNS[3]["minsize"], anchor="w")
        self.company_tree.column("district", width=COMPANY_TABLE_COLUMNS[4]["minsize"], anchor="w")

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.company_tree.yview)
        self.company_tree.configure(yscrollcommand=scrollbar.set)
        self.company_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", pady=10)

        self.company_tree.bind("<Double-1>", self.on_tree_double_click)
        self.company_tree.bind("<Button-3>", self.on_tree_right_click)

        # 초기 데이터 로드
        self.refresh_company_tree()

    # ──────────────────────────────────────
    # 업체 관리 로직
    # ──────────────────────────────────────
    def _sort_company_tree(self, col):
        """Treeview 헤더 클릭 시 오름차순/내림차순 토글 정렬"""
        ascending = not self._sort_state.get(col, False)
        self._sort_state[col] = ascending

        col_index = ["team", "company", "industry", "area", "district"].index(col)
        items = [(self.company_tree.item(iid)['values'], iid) for iid in self.company_tree.get_children()]
        items.sort(key=lambda x: str(x[0][col_index]).lower(), reverse=not ascending)

        for i, (_, iid) in enumerate(items):
            self.company_tree.move(iid, '', i)

        # 헤더 텍스트에 정렬 방향 표시
        labels = {"team": "팀장", "company": "업체명", "industry": "업종", "area": "지역", "district": "시군구"}
        for c, label in labels.items():
            arrow = ""
            if c == col:
                arrow = " ▲" if ascending else " ▼"
            self.company_tree.heading(c, text=label + arrow)

    def refresh_company_tree(self, search_text=None, force_reload=False):
        for item in self.company_tree.get_children():
            self.company_tree.delete(item)

        try:
            if not self.company_cache_loaded or force_reload:
                self.all_companies = get_all_companies()
                self.company_cache_loaded = True

            # search_text가 None이면 검색창의 현재 값을 유지
            if search_text is None:
                search_text = self.search_entry.get().strip() if hasattr(self, 'search_entry') else ""

            if search_text:
                search_lower = search_text.lower()
                companies_to_show = [
                    c for c in self.all_companies
                    if search_lower in c['team_leader'].lower() or search_lower in c['company_name'].lower()
                ]
            else:
                companies_to_show = self.all_companies

            for company in companies_to_show:
                self.company_tree.insert(
                    "", "end",
                    values=(
                        company['team_leader'], company['company_name'],
                        company['industry'], company['area'],
                        company.get('district', '전체')
                    ),
                    tags=(str(company['id']),)
                )
        except Exception as e:
            messagebox.showerror("오류", f"업체 목록 로드 실패: {e}")

    def _get_company_id_from_selection(self):
        """[Bug2 수정] 선택된 Treeview 항목에서 company_id를 int로 변환하여 반환"""
        selection = self.company_tree.selection()
        if not selection:
            return None, None
        item = self.company_tree.item(selection[0])
        try:
            company_id = int(item['tags'][0])
        except (IndexError, ValueError):
            return None, None
        company = next((c for c in self.all_companies if c['id'] == company_id), None)
        return company_id, company

    def on_tree_double_click(self, event):
        _, company = self._get_company_id_from_selection()
        if company:
            self.edit_company_dialog(company)

    def on_tree_right_click(self, event):
        # 우클릭 시 해당 행 선택
        row_id = self.company_tree.identify_row(event.y)
        if row_id:
            self.company_tree.selection_set(row_id)
        company_id, _ = self._get_company_id_from_selection()
        if company_id is not None:
            self.delete_company_confirm(company_id)

    def edit_selected_company(self):
        _, company = self._get_company_id_from_selection()
        if company is None:
            messagebox.showwarning("경고", "수정할 업체를 선택해주세요.")
            return
        self.edit_company_dialog(company)

    def delete_selected_company(self):
        company_id, _ = self._get_company_id_from_selection()
        if company_id is None:
            messagebox.showwarning("경고", "삭제할 업체를 선택해주세요.")
            return
        self.delete_company_confirm(company_id)

    def on_team_select(self, selection):
        if selection:
            companies = get_team_companies(selection)
            lines = []
            for c in companies:
                lines.append(f"  {c['company_name']}  ({c['industry']}, {c['area']})")

            display = '\n'.join(lines) if lines else '등록된 업체가 없습니다.'
            self.company_textbox.delete("0.0", "end")
            self.company_textbox.insert("0.0", display)

            if lines:
                self.status_text.set(f"{selection} 팀장의 {len(lines)}개 업체 로드 완료")
            else:
                self.status_text.set(f"{selection} 팀장의 등록된 업체가 없습니다")

    def update_district_options(self, selected_area=None):
        if not selected_area:
            selected_area = self.company_area_select.get()

        english_name = next((k for k, v in ENGLISH_AREA_CODE_MAP.items() if v == AREA_CODE_MAP.get(selected_area)), None)
        district_map_variable = f"{english_name}_DISTRICT_CODE_MAP" if english_name else None

        # code_map 모듈에서 가져오기
        import code_map
        district_code_map = getattr(code_map, district_map_variable, {"전체": "00000"}) if district_map_variable else {"전체": "00000"}

        self.company_district_combo.configure(values=list(district_code_map.keys()))
        self.company_district_combo.set("전체")

    def save_company_in_tab(self):
        team_name = self.company_team_name_entry.get().strip()
        company_name = self.company_company_name_entry.get().strip()
        industry = self.company_industry_combo.get()
        area = self.company_area_select.get()
        district = self.company_district_combo.get()

        if not all([team_name, company_name, industry, area]):
            messagebox.showwarning("경고", "모든 필드를 입력해주세요.")
            return

        if save_company_info(team_name, company_name, industry, area, district):
            messagebox.showinfo("저장 완료", "업체 정보가 저장되었습니다.")
            self.company_team_name_entry.delete(0, 'end')
            self.company_company_name_entry.delete(0, 'end')
            self.company_industry_combo.set("전체")
            self.company_area_select.set("전체")
            self.company_district_combo.set("전체")
            # [Bug1 수정] refresh_company_list_tab → refresh_company_tree
            self.refresh_company_tree(force_reload=True)
            self.load_team_leaders()
            self.status_text.set("새 업체가 성공적으로 등록되었습니다")
        else:
            messagebox.showerror("오류", "업체 정보 저장에 실패했습니다.")

    def edit_company_dialog(self, company):
        edit_window = ctk.CTkToplevel(self.root)
        edit_window.title("업체 정보 수정")
        edit_window.geometry("520x440")
        edit_window.transient(self.root)
        edit_window.grab_set()

        main_frame = ctk.CTkFrame(edit_window, fg_color=THEME["bg_main"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(
            main_frame, text="업체 정보 수정",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(pady=(10, 16))

        form = ctk.CTkFrame(main_frame, fg_color=THEME["bg_card"], corner_radius=8)
        form.pack(fill="x", padx=10, pady=(0, 16))

        grid = ctk.CTkFrame(form, fg_color="transparent")
        grid.pack(pady=16, padx=20)

        labels = ["팀장명:", "업체명:", "업종:", "지역:", "시군구:"]
        for i, label in enumerate(labels):
            ctk.CTkLabel(grid, text=label, font=make_font(12, "bold")).grid(
                row=i, column=0, sticky="w", padx=(0, 10), pady=8
            )

        edit_team = ctk.CTkEntry(grid, width=300, font=make_font(11))
        edit_team.grid(row=0, column=1, pady=8)
        edit_team.insert(0, company['team_leader'])

        edit_company = ctk.CTkEntry(grid, width=300, font=make_font(11))
        edit_company.grid(row=1, column=1, pady=8)
        edit_company.insert(0, company['company_name'])

        edit_industry = ctk.CTkComboBox(grid, values=list(INDUSTRY_CODE_MAP.keys()), width=300, font=make_font(11), state="readonly")
        edit_industry.grid(row=2, column=1, pady=8)
        edit_industry.set(company['industry'])

        edit_area = ctk.CTkComboBox(grid, values=list(AREA_CODE_MAP.keys()), width=300, font=make_font(11), state="readonly")
        edit_area.grid(row=3, column=1, pady=8)
        edit_area.set(company['area'])

        edit_district = ctk.CTkComboBox(grid, values=["전체"], width=300, font=make_font(11), state="readonly")
        edit_district.grid(row=4, column=1, pady=8)
        edit_district.set(company.get('district', '전체'))

        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(pady=8)

        def save_changes():
            new_team = edit_team.get().strip()
            new_company = edit_company.get().strip()
            new_industry = edit_industry.get()
            new_area = edit_area.get()
            new_district = edit_district.get()

            if not all([new_team, new_company, new_industry, new_area]):
                messagebox.showwarning("경고", "모든 필드를 입력해주세요.")
                return

            if update_company(company['id'], new_team, new_company, new_industry, new_area, new_district):
                messagebox.showinfo("수정 완료", "업체 정보가 수정되었습니다.")
                edit_window.destroy()
                self.refresh_company_tree(force_reload=True)
                self.load_team_leaders()
                self.status_text.set("업체 정보가 성공적으로 수정되었습니다")
            else:
                messagebox.showerror("오류", "업체 정보 수정에 실패했습니다.")

        ctk.CTkButton(
            btn_frame, text="저장", width=100, height=36,
            font=make_font(12, "bold"), fg_color=THEME["success"],
            hover_color=THEME["success_hover"], command=save_changes
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            btn_frame, text="취소", width=100, height=36,
            font=make_font(12), fg_color=THEME["secondary"],
            hover_color=THEME["secondary_hover"], command=edit_window.destroy
        ).pack(side="left", padx=8)

    def delete_company_confirm(self, company_id):
        company_id = int(company_id)
        if messagebox.askyesno("확인", "선택한 업체를 삭제하시겠습니까?"):
            if delete_company(company_id):
                messagebox.showinfo("완료", "업체가 삭제되었습니다.")
                self.refresh_company_tree(force_reload=True)
                self.load_team_leaders()
                self.status_text.set("업체가 성공적으로 삭제되었습니다")
            else:
                messagebox.showerror("오류", "업체 삭제에 실패했습니다.")

    def on_search_change(self, event=None):
        if hasattr(self, '_search_timer'):
            self.root.after_cancel(self._search_timer)

        # search_text=None으로 호출하면 검색창 값을 자동으로 읽음
        self._search_timer = self.root.after(300, lambda: self.refresh_company_tree())

    def search_companies(self):
        search_text = self.search_entry.get().strip()
        self.refresh_company_tree()
        if search_text:
            self.status_text.set(f"'{search_text}' 검색 완료")
        else:
            self.status_text.set("전체 업체 목록 표시")

    def reset_search(self):
        self.search_entry.delete(0, 'end')
        self.refresh_company_tree()
        self.status_text.set("전체 업체 목록 표시")

    # ──────────────────────────────────────
    # 공통 기능
    # ──────────────────────────────────────
    def download_data(self):
        team_leader = self.selected_team.get()
        if not team_leader:
            messagebox.showwarning("경고", "팀장을 선택해주세요.")
            return

        self.status_text.set("맞춤 공고를 생성하는 중...")
        self.root.update()
        download_data_from_db(
            team_leader=team_leader,
            log_callback=self.update_status,
            download_folder=self.settings.get("download_folder", os.getcwd())
        )

    def update_status(self, message):
        self.status_text.set(message)
        self.root.update()

    def on_tab_change(self):
        """[Bug4 수정] 탭 전환 시 항상 데이터 갱신"""
        current_tab = self.tabview.get()
        if current_tab == "업체관리":
            self.root.after(50, lambda: self.refresh_company_tree(force_reload=True))
        elif current_tab == "공고관리":
            self.load_team_leaders()
        self._update_tab_text_colors()

    def _update_tab_text_colors(self):
        """선택된 탭은 흰색, 비선택 탭은 검정색 텍스트"""
        try:
            seg = self.tabview._segmented_button
            current = seg._current_value
            for value, button in seg._buttons_dict.items():
                if value == current:
                    button.configure(text_color=THEME["white"])
                else:
                    button.configure(text_color=THEME["text_primary"])
        except Exception:
            pass

    def show_company_tab(self):
        self.tabview.set("업체관리")

    def load_initial_data(self):
        self.load_team_leaders()
        self._update_tab_text_colors()
        self.status_text.set("시스템 준비 완료 - 팀장을 선택해주세요")

    def load_team_leaders(self):
        try:
            team_leaders = get_all_team_leaders()
            self.team_combo.configure(values=team_leaders)
        except Exception as e:
            messagebox.showerror("오류", f"팀장 목록을 불러오는 중 오류 발생: {e}")

    def open_folder_settings(self):
        settings_window = ctk.CTkToplevel(self.root)
        settings_window.title("다운로드 폴더 설정")
        settings_window.geometry("520x300")
        settings_window.transient(self.root)
        settings_window.grab_set()

        main_frame = ctk.CTkFrame(settings_window, fg_color=THEME["bg_main"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(
            main_frame, text="다운로드 폴더 설정",
            font=make_font(17, "bold"), text_color=THEME["text_primary"]
        ).pack(pady=(10, 16))

        info_card = ctk.CTkFrame(main_frame, fg_color=THEME["bg_card"], corner_radius=8)
        info_card.pack(fill="x", padx=10, pady=(0, 16))

        ctk.CTkLabel(
            info_card, text="현재 다운로드 폴더:", font=make_font(12, "bold"),
            text_color=THEME["text_primary"]
        ).pack(anchor="w", padx=20, pady=(14, 4))

        current_folder = self.settings.get("download_folder", os.getcwd())
        folder_label = ctk.CTkLabel(
            info_card, text=current_folder, font=make_font(11),
            text_color=THEME["primary"], wraplength=420
        )
        folder_label.pack(anchor="w", padx=20, pady=(0, 14))

        btn_row = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_row.pack(pady=8)

        def select_folder():
            selected = filedialog.askdirectory(
                title="다운로드 폴더 선택",
                initialdir=self.settings.get("download_folder", os.getcwd())
            )
            if selected:
                self.settings["download_folder"] = selected
                if save_settings(self.settings):
                    folder_label.configure(text=selected)
                    self._update_folder_display()
                    messagebox.showinfo("완료", "다운로드 폴더가 설정되었습니다.")

        def reset_folder():
            self.settings["download_folder"] = os.getcwd()
            if save_settings(self.settings):
                folder_label.configure(text=os.getcwd())
                self._update_folder_display()
                messagebox.showinfo("완료", "기본 폴더로 복원되었습니다.")

        ctk.CTkButton(
            btn_row, text="폴더 선택", font=make_font(13, "bold"),
            width=130, height=38, fg_color=THEME["success"],
            hover_color=THEME["success_hover"], command=select_folder
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            btn_row, text="기본값 복원", font=make_font(12),
            width=110, height=34, fg_color=THEME["warning"],
            hover_color="#e0a800", command=reset_folder
        ).pack(side="left", padx=8)

        ctk.CTkButton(
            btn_row, text="닫기", font=make_font(12),
            width=80, height=34, fg_color=THEME["secondary"],
            hover_color=THEME["secondary_hover"], command=settings_window.destroy
        ).pack(side="left", padx=8)

    def _update_folder_display(self):
        current_folder = self.settings.get("download_folder", os.getcwd())
        folder_display = current_folder if len(current_folder) < 60 else "..." + current_folder[-57:]
        if hasattr(self, 'folder_display_label'):
            self.folder_display_label.configure(text=f"다운로드 폴더: {folder_display}")

    def _card(self, parent):
        """모던한 카드 UI 생성"""
        return ctk.CTkFrame(parent, fg_color=THEME["bg_card"], corner_radius=10)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = ModernTeamManagerApp()
    app.run()
